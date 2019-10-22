import datetime
import calendar
import openpyxl
import os

weekly = ("Mon", "Tue", "Wed", "Thur", "Fri", "Sat", "Sun")

input_date = input('Input a target year and month[ex: 201807] : ')

month_current = datetime.datetime.strptime(input_date + '01', '%Y%m%d')

# month_last = month_current - datetime.timedelta(days=1)

date_start = month_current.replace(day=1)
_, lastday = calendar.monthrange(month_current.year, month_current.month)
date_end = month_current.replace(day=lastday)

current_dir = os.path.dirname(__file__)
file_template = 'attendance_management.xlsx'
file_path = current_dir + '/' + file_template
wb = openpyxl.load_workbook(file_path)

ws = wb['Sheet1']

ws.cell(row=1, column=1).value = month_current.year
ws.cell(row=1, column=3).value = month_current.month

for row_num in range(13, 44):
    if date_start <= date_end:
        ws.cell(row=row_num, column=1).value = date_start.month
        ws.cell(row=row_num, column=2).value = date_start.day
        ws.cell(row=row_num, column=3).value = weekly[date_start.weekday()]
    date_start += datetime.timedelta(days=1)

file_new = current_dir + '/' + \
    month_current.strftime('%Y%m') + '_' + file_template
wb.save(file_new)
